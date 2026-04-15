"""
Export helpers — convert a filtered (CanonicalClaim, ValidationResult) list
to downloadable CSV or Excel bytes.

Both functions are pure Python (no Streamlit imports) so they can be tested
independently and reused outside the UI layer.
"""

from __future__ import annotations

import csv
import io
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from utils.claim_display import patient_name, dos

if TYPE_CHECKING:
    from parser.models import CanonicalClaim
    from validator.snip import ValidationResult

# ---------------------------------------------------------------------------
# Shared column definitions
# ---------------------------------------------------------------------------

_HEADERS = [
    "Claim ID",
    "Billing NPI",
    "Patient Name",
    "Date of Service",
    "Total Charge",
    "Status",
    "Error Count",
    "Warning Count",
    "File Name",
]

_COL_WIDTHS = [20, 15, 28, 22, 14, 10, 13, 14, 30]

# Excel fill colours (no '#' prefix — openpyxl requires plain hex)
_FILL_HEADER = PatternFill("solid", fgColor="1F3864")   # dark navy
_FILL_PASS   = PatternFill("solid", fgColor="C6EFCE")   # Excel-style green
_FILL_FAIL   = PatternFill("solid", fgColor="FFC7CE")   # Excel-style red
_FONT_HEADER = Font(bold=True, color="FFFFFF")


def _row_values(
    canonical: "CanonicalClaim",
    result: "ValidationResult",
) -> list:
    """Extract a single export row as a plain list."""
    c = canonical.claim
    err_n  = sum(1 for e in result.errors if e.severity == "error")
    warn_n = sum(1 for e in result.errors if e.severity == "warning")
    return [
        c.claim_id or "",
        c.billing_provider.npi or "",
        patient_name(canonical),
        dos(canonical),
        float(c.total_charge),
        result.status,
        err_n,
        warn_n,
        canonical.file.file_name if canonical.file else "",
    ]


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def build_csv(
    pairs: list[tuple["CanonicalClaim", "ValidationResult"]],
) -> bytes:
    """
    Serialise *pairs* to UTF-8-sig encoded CSV bytes.

    The BOM prefix (\\xef\\xbb\\xbf) ensures Microsoft Excel opens the file
    without garbling special characters on double-click.

    Parameters
    ----------
    pairs:
        Filtered list of (CanonicalClaim, ValidationResult).  An empty list
        produces a header-only CSV.

    Returns
    -------
    bytes
        Ready to pass directly to ``st.download_button(data=...)``.
    """
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_HEADERS)
    for canonical, result in pairs:
        writer.writerow(_row_values(canonical, result))
    return buf.getvalue().encode("utf-8-sig")


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def build_excel(
    pairs: list[tuple["CanonicalClaim", "ValidationResult"]],
) -> bytes:
    """
    Serialise *pairs* to an .xlsx workbook as bytes.

    Formatting:
    - Header row: bold white text on dark-navy background, frozen.
    - Pass rows: light-green fill (Excel traffic-light convention).
    - Fail rows: light-red fill.
    - Column widths are pre-set for readability.

    Parameters
    ----------
    pairs:
        Filtered list of (CanonicalClaim, ValidationResult).  An empty list
        produces a header-only sheet.

    Returns
    -------
    bytes
        Ready to pass directly to ``st.download_button(data=...)``.

    Note
    ----
    openpyxl loads the entire workbook in memory.  For very large exports
    (tens of thousands of rows) consider switching to write_only=True mode.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Claims"

    # ── Header row ───────────────────────────────────────────────────────────
    ws.append(_HEADERS)
    for col_idx, (cell, width) in enumerate(
        zip(ws[1], _COL_WIDTHS), start=1
    ):
        cell.font       = _FONT_HEADER
        cell.fill       = _FILL_HEADER
        cell.alignment  = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze the header row so it stays visible while scrolling
    ws.freeze_panes = "A2"

    # ── Data rows ─────────────────────────────────────────────────────────────
    for canonical, result in pairs:
        row_values = _row_values(canonical, result)
        ws.append(row_values)

        fill = _FILL_PASS if result.status == "Pass" else _FILL_FAIL
        row_idx = ws.max_row
        for col_idx in range(1, len(_HEADERS) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    # ── Serialise to bytes ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
