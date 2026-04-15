"""
Tests for the export layer (ui/export.py).

All tests are pure Python — no Streamlit required.
"""

from __future__ import annotations

import csv
import io
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from parser.models import (
    BillingProvider, CanonicalClaim, Claim, FileEnvelope,
    ServiceLine, Subscriber, TransactionEnvelope,
)
from validator.snip import ValidationError, ValidationResult
from ui.export import build_csv, build_excel, _HEADERS


# ---------------------------------------------------------------------------
# Shared fixture helpers
# ---------------------------------------------------------------------------

def _make_pair(
    claim_id: str = "CLM001",
    status: str = "Pass",
    npi: str = "1234567890",
    charge: Decimal = Decimal("250.00"),
    payer: str = "BlueCross",
    file_name: str = "test.edi",
    errors: list | None = None,
) -> tuple[CanonicalClaim, ValidationResult]:
    canonical = CanonicalClaim(
        file=FileEnvelope(
            file_name=file_name,
            sender_id="SENDER01",
            receiver_id="RECEIVER01",
        ),
        transaction=TransactionEnvelope(st_control_number="0001"),
        claim=Claim(
            claim_id=claim_id,
            total_charge=charge,
            billing_provider=BillingProvider(npi=npi),
            subscriber=Subscriber(
                member_id="MEM001",
                last_name="Smith",
                first_name="Jane",
                payer_name=payer,
            ),
            service_lines=[
                ServiceLine(
                    line_number=1,
                    charge=charge,
                    date="2024-03-01",
                    procedure_code="99213",
                ),
            ],
        ),
    )
    result = ValidationResult(
        claim_id=claim_id,
        status=status,
        errors=errors or [],
    )
    return canonical, result


def _make_error(severity: str = "error", level: int = 2, code: str = "L2-MISSING-HI") -> ValidationError:
    return ValidationError(
        level=level,
        severity=severity,
        code=code,
        message="Test validation error",
        loop="2300",
        position=5,
        raw_segment="CLM*CLM001*250",
    )


# ---------------------------------------------------------------------------
# CSV tests
# ---------------------------------------------------------------------------

class TestBuildCsv:
    def test_returns_bytes(self):
        result = build_csv([_make_pair()])
        assert isinstance(result, bytes)

    def test_utf8_bom_prefix(self):
        result = build_csv([_make_pair()])
        assert result[:3] == b"\xef\xbb\xbf"

    def test_header_row_contains_all_columns(self):
        result = build_csv([])
        text = result.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        header = next(reader)
        for col in _HEADERS:
            assert col in header, f"Missing column: {col}"

    def test_empty_pairs_produces_header_only(self):
        result = build_csv([])
        text = result.decode("utf-8-sig")
        rows = [r for r in csv.reader(io.StringIO(text)) if r]
        assert len(rows) == 1  # header only

    def test_row_count_matches_pairs(self):
        pairs = [_make_pair(claim_id=f"CLM{i:03d}") for i in range(5)]
        result = build_csv(pairs)
        text = result.decode("utf-8-sig")
        rows = [r for r in csv.reader(io.StringIO(text)) if r]
        assert len(rows) == 6  # 1 header + 5 data

    def test_claim_id_in_output(self):
        result = build_csv([_make_pair(claim_id="SPECIAL_CLM")])
        assert b"SPECIAL_CLM" in result

    def test_pass_status_in_output(self):
        result = build_csv([_make_pair(status="Pass")])
        assert b"Pass" in result

    def test_fail_status_in_output(self):
        result = build_csv([_make_pair(status="Fail")])
        assert b"Fail" in result

    def test_error_and_warning_counts(self):
        errors = [
            _make_error(severity="error"),
            _make_error(severity="error"),
            _make_error(severity="warning"),
        ]
        pair = _make_pair(status="Fail", errors=errors)
        result = build_csv([pair])
        text = result.decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(text)))
        data_row = rows[1]
        # Error Count column (index 6) and Warning Count (index 7)
        assert data_row[6] == "2"
        assert data_row[7] == "1"

    def test_file_name_in_output(self):
        result = build_csv([_make_pair(file_name="batch_jan.edi")])
        assert b"batch_jan.edi" in result


# ---------------------------------------------------------------------------
# Excel tests
# ---------------------------------------------------------------------------

class TestBuildExcel:
    def test_returns_bytes(self):
        result = build_excel([_make_pair()])
        assert isinstance(result, bytes)

    def test_valid_xlsx_magic_bytes(self):
        """OOXML (.xlsx) files are ZIP archives — first 4 bytes are PK\\x03\\x04."""
        result = build_excel([_make_pair()])
        assert result[:4] == b"PK\x03\x04"

    def test_header_row_is_bold(self):
        result = build_excel([_make_pair()])
        ws = load_workbook(io.BytesIO(result)).active
        assert ws.cell(row=1, column=1).font.bold is True

    def test_header_cells_have_all_columns(self):
        result = build_excel([])
        ws = load_workbook(io.BytesIO(result)).active
        header_vals = [ws.cell(row=1, column=c).value for c in range(1, len(_HEADERS) + 1)]
        assert header_vals == _HEADERS

    def test_freeze_pane_at_a2(self):
        result = build_excel([_make_pair()])
        ws = load_workbook(io.BytesIO(result)).active
        assert ws.freeze_panes == "A2"

    def test_pass_row_green_fill(self):
        result = build_excel([_make_pair(status="Pass")])
        ws = load_workbook(io.BytesIO(result)).active
        fill = ws.cell(row=2, column=1).fill
        # openpyxl prefixes the color with alpha — ends with the hex we set
        assert fill.fgColor.rgb.endswith("C6EFCE")

    def test_fail_row_red_fill(self):
        result = build_excel([_make_pair(status="Fail")])
        ws = load_workbook(io.BytesIO(result)).active
        fill = ws.cell(row=2, column=1).fill
        assert fill.fgColor.rgb.endswith("FFC7CE")

    def test_row_count_matches_pairs(self):
        pairs = [_make_pair(claim_id=f"CLM{i:03d}") for i in range(4)]
        result = build_excel(pairs)
        ws = load_workbook(io.BytesIO(result)).active
        # max_row includes header
        assert ws.max_row == 5

    def test_claim_id_in_cell_a2(self):
        result = build_excel([_make_pair(claim_id="CHECK001")])
        ws = load_workbook(io.BytesIO(result)).active
        assert ws.cell(row=2, column=1).value == "CHECK001"

    def test_empty_pairs_produces_header_only(self):
        result = build_excel([])
        ws = load_workbook(io.BytesIO(result)).active
        assert ws.max_row == 1

    def test_charge_stored_as_numeric(self):
        """Total Charge column must be a number, not a string."""
        result = build_excel([_make_pair(charge=Decimal("123.45"))])
        ws = load_workbook(io.BytesIO(result)).active
        charge_val = ws.cell(row=2, column=5).value
        assert isinstance(charge_val, (int, float))
        assert abs(charge_val - 123.45) < 0.001
